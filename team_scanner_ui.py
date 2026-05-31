import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox
import threading
import openpyxl
from duckduckgo_search import DDGS
import requests
from bs4 import BeautifulSoup
import re
import time
import sys
import urllib3

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

FILE_PATH = r'C:\Coach-C\agents.xlsx'

IGNORED_DOMAINS = [
    "remax.", "centris.ca", "realtor.ca", "facebook.com", "linkedin.com",
    "instagram.com", "rate-my-agent", "agentpronto", "royallepage",
    "century21", "kijiji", "yellowpages", "twitter.com", "youtube.com",
    "tiktok.com", "proprio-direct", "sutton", "kw.com", "kellerwilliams",
    "groupeviacapitale", "via-capitale", "exprealty", "exp-realty"
]

TEAM_KEYWORDS = [r'\bteam\b', r'\bgroup\b', r'\bgroupe\b', r'\bequipe\b', r'\béquipe\b']
keyword_pattern = re.compile('|'.join(TEAM_KEYWORDS), re.IGNORECASE)

class ScannerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Team Scanner Batch Tool")
        self.root.geometry("650x550")
        
        # Load sheets
        self.sheets = []
        try:
            wb = openpyxl.load_workbook(FILE_PATH, read_only=True)
            self.sheets = wb.sheetnames
            wb.close()
        except PermissionError:
            messagebox.showerror("Error", f"agents.xlsx is currently open in Excel. Please close it before starting the tool.")
            sys.exit(1)
        except Exception as e:
            messagebox.showerror("Error", f"Could not load {FILE_PATH}.\n\nError: {e}")
            sys.exit(1)
            
        ttk.Label(root, text="1. Select Province Tab:", font=("Arial", 10, "bold")).pack(pady=(15, 5))
        self.sheet_var = tk.StringVar(value=self.sheets[0] if self.sheets else "")
        self.dropdown = ttk.Combobox(root, textvariable=self.sheet_var, values=self.sheets, state="readonly", width=30)
        self.dropdown.pack(pady=5)
        
        self.start_btn = ttk.Button(root, text="Start Scanning Batch (50 Agents)", command=self.start_batch)
        self.start_btn.pack(pady=15)
        
        ttk.Label(root, text="Progress Log:", font=("Arial", 10)).pack(anchor="w", padx=20)
        self.log = scrolledtext.ScrolledText(root, width=75, height=20, state='disabled', font=("Consolas", 9))
        self.log.pack(pady=5, padx=20)
        
        self.is_running = False

    def log_msg(self, msg):
        self.log.config(state='normal')
        self.log.insert(tk.END, msg + "\n")
        self.log.see(tk.END)
        self.log.config(state='disabled')
        self.root.update()

    def start_batch(self):
        if self.is_running:
            return
        sheet_name = self.sheet_var.get()
        if not sheet_name:
            return
            
        self.is_running = True
        self.start_btn.config(state='disabled')
        self.log.config(state='normal')
        self.log.delete(1.0, tk.END)
        self.log.config(state='disabled')
        
        threading.Thread(target=self.run_scan, args=(sheet_name,), daemon=True).start()

    def run_scan(self, sheet_name):
        self.log_msg(f"Starting batch scan for tab: {sheet_name}")
        self.log_msg("Loading spreadsheet (make sure it is closed in Excel)...")
        
        try:
            wb = openpyxl.load_workbook(FILE_PATH)
            sheet = wb[sheet_name]
            
            headers = [cell.value for cell in sheet[1]]
            col_map = {name: idx+1 for idx, name in enumerate(headers) if name}
            max_col = sheet.max_column
            
            for new_col in ['Website Scanned', 'Verified Team']:
                if new_col not in col_map:
                    max_col += 1
                    sheet.cell(row=1, column=max_col, value=new_col)
                    col_map[new_col] = max_col
            
            agents_to_scan = []
            for row in range(2, sheet.max_row + 1):
                agent_name = sheet.cell(row=row, column=1).value
                if not agent_name: continue
                
                scanned_status = sheet.cell(row=row, column=col_map['Website Scanned']).value
                if not scanned_status:
                    city = sheet.cell(row=row, column=col_map.get('City', 4)).value or ""
                    agents_to_scan.append((row, agent_name, city))
                    
                if len(agents_to_scan) >= 50:
                    break
                    
            if not agents_to_scan:
                self.log_msg("No unscanned agents found in this tab! You are all caught up.")
                wb.close()
                self.finish_scan()
                return
                
            self.log_msg(f"Found {len(agents_to_scan)} agents to scan. Connecting to DuckDuckGo...\n")
            
            ddgs = DDGS()
            req_headers = {'User-Agent': 'Mozilla/5.0'}
            
            for i, (row, agent_name, city) in enumerate(agents_to_scan, 1):
                if not self.is_running: break
                
                self.log_msg(f"[{i}/{len(agents_to_scan)}] Scanning: {agent_name} ({city})")
                
                query = f'"{agent_name}" real estate courtier {city}'
                custom_url = None
                
                try:
                    results = ddgs.text(query, max_results=5)
                    if results:
                        for r in results:
                            url = r.get('href', '').lower()
                            if not any(ign in url for ign in IGNORED_DOMAINS):
                                custom_url = r.get('href')
                                break
                except Exception as e:
                    self.log_msg(f"   -> Search error: {e}")
                    
                if custom_url:
                    self.log_msg(f"   -> Testing URL: {custom_url}")
                    try:
                        resp = requests.get(custom_url, headers=req_headers, timeout=8, verify=False)
                        soup = BeautifulSoup(resp.text, 'html.parser')
                        page_text = soup.get_text(" ", strip=True)
                        
                        if keyword_pattern.search(page_text):
                            self.log_msg("   -> *** VERIFIED TEAM ***")
                            sheet.cell(row=row, column=col_map['Verified Team'], value="Yes")
                        else:
                            self.log_msg("   -> No team keywords found.")
                            sheet.cell(row=row, column=col_map['Verified Team'], value="No")
                            
                        sheet.cell(row=row, column=col_map['Website Scanned'], value=custom_url)
                        
                    except Exception as e:
                        self.log_msg(f"   -> Failed to load website.")
                        sheet.cell(row=row, column=col_map['Website Scanned'], value="Failed Load")
                else:
                    self.log_msg("   -> No custom domain found.")
                    sheet.cell(row=row, column=col_map['Website Scanned'], value="No Website")
                    
                if i % 5 == 0:
                    try:
                        wb.save(FILE_PATH)
                    except:
                        pass
                
                time.sleep(2)
                
            if sheet.tables:
                table = list(sheet.tables.values())[0]
                existing_cols = len(table.tableColumns)
                
                if existing_cols < max_col:
                    from openpyxl.worksheet.table import TableColumn
                    for col_idx in range(existing_cols + 1, max_col + 1):
                        header_name = str(sheet.cell(row=1, column=col_idx).value)
                        new_col = TableColumn(id=col_idx, name=header_name)
                        table.tableColumns.append(new_col)
                        
                table.ref = f"A1:{openpyxl.utils.get_column_letter(max_col)}{sheet.max_row}"
                
            wb.save(FILE_PATH)
            self.log_msg("\nBatch complete! Spreadsheet updated.")
            self.log_msg("You can now close this window or click Start again for the next 50.")
            
        except PermissionError:
            self.log_msg("\nERROR: Excel file is currently open. Please close agents.xlsx and try again.")
        except Exception as e:
            self.log_msg(f"\nERROR: {str(e)}")
            
        self.finish_scan()

    def finish_scan(self):
        self.is_running = False
        self.start_btn.config(state='normal')

if __name__ == "__main__":
    root = tk.Tk()
    app = ScannerApp(root)
    root.mainloop()
