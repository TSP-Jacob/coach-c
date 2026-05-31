import openpyxl
from openpyxl.worksheet.table import TableColumn
import sys

file_path = 'agents.xlsx'

try:
    with open(file_path, 'a'):
        pass
except PermissionError:
    print("PERMISSION_ERROR")
    sys.exit(1)

wb = openpyxl.load_workbook(file_path)
sheet = wb['Quebec']

if sheet.tables:
    table = list(sheet.tables.values())[0]
    existing_cols = len(table.tableColumns)
    max_col = sheet.max_column
    
    if existing_cols < max_col:
        for col_idx in range(existing_cols + 1, max_col + 1):
            header_name = str(sheet.cell(row=1, column=col_idx).value)
            new_col = TableColumn(id=col_idx, name=header_name)
            table.tableColumns.append(new_col)
            
        table.ref = f"A1:{openpyxl.utils.get_column_letter(max_col)}{sheet.max_row}"
        wb.save(file_path)
        print("FIXED")
    else:
        print("ALREADY FIXED")
