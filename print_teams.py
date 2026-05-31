import openpyxl

wb = openpyxl.load_workbook('agents.xlsx')
target_sheet = wb['Target Teams']

for row in range(2, target_sheet.max_row + 1):
    print(target_sheet.cell(row=row, column=1).value)
