import requests
from bs4 import BeautifulSoup
import re
import time
import openpyxl
import sys

file_path = 'agents.xlsx'

try:
    with open(file_path, 'a'):
        pass
except PermissionError:
    print("PERMISSION_ERROR")
    sys.exit(1)

base_url = "https://www.rate-my-agent.com/Quebec-QC-Agent-Ratings"
headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
    'Accept-Language': 'en-US,en;q=0.9',
}

agents = []
page = 1

while True:
    print(f"Fetching page {page}...")
    url = f"{base_url}?page={page}" if page > 1 else base_url
    r = requests.get(url, headers=headers)
    
    if r.status_code != 200:
        print(f"Failed to fetch page {page}, status code {r.status_code}")
        break
        
    soup = BeautifulSoup(r.text, 'html.parser')
    
    page_agents = 0
    seen_pagination = False
    
    for a in soup.find_all('a'):
        href = a.get('href', '')
        text = a.get_text(strip=True)
        
        if '?page=' in href:
            seen_pagination = True
            
        if '-ratings-' in href and not seen_pagination:
            if text and text != '...':
                if text not in agents:
                    agents.append(text)
                    page_agents += 1
                    
    print(f"Found {page_agents} agents on page {page}.")
    
    if page_agents == 0 and page > 1:
        break
        
    next_page_link = soup.find('a', href=re.compile(rf'\?page={page+1}'))
    if not next_page_link:
        break
        
    page += 1
    time.sleep(1)

print(f"Found {len(agents)} agents in total.")

wb = openpyxl.load_workbook(file_path)

if 'Quebec' not in wb.sheetnames:
    print("Creating 'Quebec' tab...")
    sheet = wb.create_sheet('Quebec')
else:
    sheet = wb['Quebec']

if sheet['A1'].value is None:
    sheet['A1'] = 'Agent Name'

for i, agent in enumerate(agents, start=2):
    sheet[f'A{i}'] = agent

wb.save(file_path)
print("SUCCESS")
