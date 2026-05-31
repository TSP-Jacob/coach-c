import openpyxl
import sys
import re

file_path = 'agents.xlsx'

try:
    with open(file_path, 'a'):
        pass
except PermissionError:
    print("PERMISSION_ERROR")
    sys.exit(1)

wb = openpyxl.load_workbook(file_path)

if 'Quebec' not in wb.sheetnames:
    print("Error: Quebec tab not found.")
    sys.exit(1)

source_sheet = wb['Quebec']

# Create or clear Target Teams tab
target_sheet_name = 'Target Teams'
if target_sheet_name in wb.sheetnames:
    del wb[target_sheet_name]
target_sheet = wb.create_sheet(target_sheet_name)

# Copy headers
headers = [cell.value for cell in source_sheet[1]]
for col_idx, header in enumerate(headers, start=1):
    target_sheet.cell(row=1, column=col_idx, value=header)

# Keywords to identify teams
keywords = [r'\bteam\b', r'\bgroup\b', r'\bgroupe\b', r'\bequipe\b', r'\béquipe\b']
pattern = re.compile('|'.join(keywords), re.IGNORECASE)

team_count = 0
next_row = 2

for row in range(2, source_sheet.max_row + 1):
    agent_name = source_sheet.cell(row=row, column=1).value
    if agent_name:
        name_str = str(agent_name)
        if pattern.search(name_str):
            # It's a team! Copy the row
            for col_idx in range(1, source_sheet.max_column + 1):
                target_sheet.cell(row=next_row, column=col_idx, value=source_sheet.cell(row=row, column=col_idx).value)
            next_row += 1
            team_count += 1

# Format as table
if team_count > 0:
    from openpyxl.worksheet.table import Table, TableStyleInfo
    table_ref = f"A1:{openpyxl.utils.get_column_letter(source_sheet.max_column)}{next_row - 1}"
    tab = Table(displayName="TeamsTable", ref=table_ref)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    target_sheet.add_table(tab)

wb.save(file_path)
print(f"SUCCESS: {team_count} teams found.")
