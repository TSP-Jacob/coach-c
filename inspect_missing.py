import openpyxl

wb = openpyxl.load_workbook('agents.xlsx')
sheet = wb['Quebec']

missing_metrics = 0
total_agents = 0
names_missing = []

for row in range(2, sheet.max_row + 1):
    agent_name = sheet.cell(row=row, column=1).value
    if not agent_name: continue
    total_agents += 1
    
    # Check if Recent Sales, Years Exp, and Total Reviews are all N/A
    sales = sheet.cell(row=row, column=5).value  # E
    exp = sheet.cell(row=row, column=6).value    # F
    rev = sheet.cell(row=row, column=7).value    # G
    
    if sales == "N/A" and exp == "N/A" and rev == "N/A":
        missing_metrics += 1
        if len(names_missing) < 10:
            names_missing.append(agent_name)

print(f"Total Agents: {total_agents}")
print(f"Agents with NO metrics: {missing_metrics}")
print(f"Sample missing names: {names_missing}")

if sheet.tables:
    table = list(sheet.tables.values())[0]
    print(f"Table cols count: {len(table.tableColumns)}, max_col: {sheet.max_column}")
