import requests
from bs4 import BeautifulSoup
import re
import time
import openpyxl
import sys

file_path = 'agents.xlsx'

try:
    with open(file_path, 'a'):
        pass
except PermissionError:
    print("PERMISSION_ERROR")
    sys.exit(1)

wb = openpyxl.load_workbook(file_path)
sheet = wb['Quebec']

city_col_idx = 4
for cell in sheet[1]:
    if cell.value == 'City':
        city_col_idx = cell.column
        break

seen_names = set()
for row in range(2, sheet.max_row + 1):
    agent_name = sheet.cell(row=row, column=1).value
    if agent_name:
        seen_names.add(str(agent_name).strip().lower())

next_row = sheet.max_row + 1

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
    'Accept-Language': 'en-US,en;q=0.9',
}

urls_and_cities = [
    ("Gatineau", "https://www.rate-my-agent.com/Gatineau-QC-Agent-Ratings"),
    ("Laval", "https://www.rate-my-agent.com/Laval-QC-Real-Estate-Agent-Reviews-Ratings"),
    ("Longueuil", "https://www.rate-my-agent.com/Longueuil-QC-Real-Estate-Agent-Reviews-Ratings"),
    ("Sherbrooke", "https://www.rate-my-agent.com/Sherbrooke-QC-Real-Estate-Agent-Reviews-Ratings"),
    ("Levis", "https://www.rate-my-agent.com/Levis-QC-Real-Estate-Agent-Reviews-Ratings")
]

total_added = 0

for city, base_url in urls_and_cities:
    page = 1
    print(f"\n--- Scraping {city} ---")
    
    while True:
        print(f"Fetching {city} page {page}...")
        url = f"{base_url}?page={page}" if page > 1 else base_url
        r = requests.get(url, headers=headers)
        
        if r.status_code != 200:
            print(f"Failed to fetch page {page}, status code {r.status_code}")
            break
            
        soup = BeautifulSoup(r.text, 'html.parser')
        
        page_agents = 0
        seen_pagination = False
        
        for a in soup.find_all('a'):
            href = a.get('href', '')
            text = a.get_text(strip=True)
            
            if '?page=' in href:
                seen_pagination = True
                
            if '-ratings-' in href and not seen_pagination:
                if text and text != '...':
                    clean_name = text.strip()
                    lower_name = clean_name.lower()
                    
                    if lower_name not in seen_names:
                        seen_names.add(lower_name)
                        sheet.cell(row=next_row, column=1, value=clean_name)
                        sheet.cell(row=next_row, column=city_col_idx, value=city)
                        next_row += 1
                        page_agents += 1
                        total_added += 1
                        
        print(f"Added {page_agents} new agents from {city} page {page}.")
        
        # We don't break if page_agents == 0 here because maybe a page had all duplicates
        # We only break if there's no next page link
            
        next_page_link = soup.find('a', href=re.compile(rf'\?page={page+1}'))
        if not next_page_link:
            break
            
        page += 1
        time.sleep(1)

print(f"\nTotal new agents added across all cities: {total_added}")

if sheet.tables:
    table = list(sheet.tables.values())[0]
    table.ref = f"A1:K{next_row - 1}"

wb.save(file_path)
print("SUCCESS")
