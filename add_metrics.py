import requests
from bs4 import BeautifulSoup
import re
import time
import openpyxl
import sys

file_path = 'agents.xlsx'

try:
    with open(file_path, 'a'):
        pass
except PermissionError:
    print("PERMISSION_ERROR")
    sys.exit(1)

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
    'Accept-Language': 'en-US,en;q=0.9',
}

agent_metrics = {}

pronto_urls = [
    "https://agentpronto.com/quebec/montreal",
    "https://agentpronto.com/quebec/laval"
]

print("Scraping Agent Pronto...")
for base_url in pronto_urls:
    page = 1
    while True:
        url = f"{base_url}?page={page}" if page > 1 else base_url
        r = requests.get(url, headers=headers)
        if r.status_code != 200: break
        soup = BeautifulSoup(r.text, 'html.parser')
        links = soup.find_all('a', href=re.compile(r'^/agents/'))
        if not links: break
        for link in links:
            name_span = link.find('span', class_=re.compile(r'font-extrabold'))
            if not name_span: continue
            name = name_span.get_text(strip=True).lower()
            text = link.get_text(" ", strip=True)
            sales = int((re.search(r'(\d+)\s+recent sales', text) or [0,0])[1])
            exp = int((re.search(r'(\d+)\s+years experience', text) or [0,0])[1])
            rev = int((re.search(r'\((\d+)\)', text) or [0,0])[1])
            agent_metrics[name] = {'sales': sales, 'exp': exp, 'reviews': rev}
        
        next_page = soup.find('a', href=re.compile(rf'\?page={page+1}'))
        if not next_page: break
        page += 1
        time.sleep(0.5)

rma_urls = [
    "https://www.rate-my-agent.com/Quebec-QC-Agent-Ratings",
    "https://www.rate-my-agent.com/Gatineau-QC-Agent-Ratings",
    "https://www.rate-my-agent.com/Laval-QC-Real-Estate-Agent-Reviews-Ratings",
    "https://www.rate-my-agent.com/Longueuil-QC-Real-Estate-Agent-Reviews-Ratings",
    "https://www.rate-my-agent.com/Sherbrooke-QC-Real-Estate-Agent-Reviews-Ratings",
    "https://www.rate-my-agent.com/Levis-QC-Real-Estate-Agent-Reviews-Ratings"
]

print("Scraping Rate-My-Agent...")
for base_url in rma_urls:
    page = 1
    while True:
        url = f"{base_url}?page={page}" if page > 1 else base_url
        r = requests.get(url, headers=headers)
        if r.status_code != 200: break
        soup = BeautifulSoup(r.text, 'html.parser')
        
        page_agents = 0
        seen_pagination = False
        for a in soup.find_all('a'):
            href = a.get('href', '')
            text = a.get_text(strip=True)
            if '?page=' in href: seen_pagination = True
            if '-ratings-' in href and not seen_pagination and text and text != '...':
                name = text.strip().lower()
                parent = a.find_parent('div', class_='row') or a.find_parent('div')
                ptext = parent.get_text(" ", strip=True) if parent else ""
                total = int((re.search(r'(\d+)\s+total', ptext) or [0,0])[1])
                if name not in agent_metrics:
                    agent_metrics[name] = {'sales': 0, 'exp': 0, 'reviews': total}
                elif agent_metrics[name]['reviews'] == 0:
                    agent_metrics[name]['reviews'] = total
                page_agents += 1
                
        if page_agents == 0 and page > 1: break
        next_page = soup.find('a', href=re.compile(rf'\?page={page+1}'))
        if not next_page: break
        page += 1
        time.sleep(0.5)

print("Updating Excel...")
wb = openpyxl.load_workbook(file_path)
sheet = wb['Quebec']

# Find header row and column indices
headers_row = 1
max_col = sheet.max_column
col_map = {}
for cell in sheet[headers_row]:
    col_map[cell.value] = cell.column

# Add new columns if not exist
for h in ['Recent Sales', 'Years Experience', 'Total Reviews', 'Activity Rating']:
    if h not in col_map:
        max_col += 1
        sheet.cell(row=headers_row, column=max_col, value=h)
        col_map[h] = max_col

for row in range(2, sheet.max_row + 1):
    agent_name = sheet.cell(row=row, column=1).value
    if not agent_name: continue
    lower_name = str(agent_name).strip().lower()
    
    metrics = agent_metrics.get(lower_name, {'sales': 0, 'exp': 0, 'reviews': 0})
    sales = metrics['sales']
    exp = metrics['exp']
    rev = metrics['reviews']
    
    sheet.cell(row=row, column=col_map['Recent Sales'], value=sales if sales > 0 else "N/A")
    sheet.cell(row=row, column=col_map['Years Experience'], value=exp if exp > 0 else "N/A")
    sheet.cell(row=row, column=col_map['Total Reviews'], value=rev if rev > 0 else "N/A")
    
    # Calculate Activity Rating (1-5)
    score = 1
    if sales > 0:
        if sales >= 151: score = 5
        elif sales >= 81: score = 4
        elif sales >= 31: score = 3
        elif sales >= 11: score = 2
    else:
        if rev >= 101: score = 5
        elif rev >= 31: score = 4
        elif rev >= 11: score = 3
        elif rev >= 3: score = 2
        
    # Team boost
    if any(word in lower_name for word in ['team', 'group', 'equipe', 'équipe']):
        score = min(5, score + 1)
        
    sheet.cell(row=row, column=col_map['Activity Rating'], value=score)

if sheet.tables:
    table = list(sheet.tables.values())[0]
    table.ref = f"A1:{openpyxl.utils.get_column_letter(max_col)}{sheet.max_row}"

wb.save(file_path)
print("SUCCESS")
