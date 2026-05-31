import requests
from bs4 import BeautifulSoup
import re
import time
import openpyxl
import sys

file_path = 'agents.xlsx'

try:
    with open(file_path, 'a'):
        pass
except PermissionError:
    print("PERMISSION_ERROR")
    sys.exit(1)

base_url = "https://agentpronto.com/quebec/laval"
headers = {'User-Agent': 'Mozilla/5.0'}
agents = []

page = 1
while True:
    print(f"Fetching page {page}...")
    url = f"{base_url}?page={page}" if page > 1 else base_url
    r = requests.get(url, headers=headers)
    if r.status_code != 200:
        break
        
    soup = BeautifulSoup(r.text, 'html.parser')
    agent_links = soup.find_all('a', href=re.compile(r'^/agents/'))
    if not agent_links:
        break
        
    for link in agent_links:
        name_span = link.find('span', class_=re.compile(r'font-extrabold'))
        if name_span:
            name = name_span.get_text(strip=True)
            if name and name not in agents:
                agents.append(name)
                
    next_page = soup.find('a', href=re.compile(rf'page={page+1}'))
    if not next_page:
        break
        
    page += 1
    time.sleep(0.5)

print(f"Found {len(agents)} agents in Laval.")

wb = openpyxl.load_workbook(file_path)

if 'Laval' not in wb.sheetnames:
    print("Creating 'Laval' tab...")
    sheet = wb.create_sheet('Laval')
else:
    sheet = wb['Laval']

if sheet['A1'].value is None:
    sheet['A1'] = 'Agent Name'

for i, agent in enumerate(agents, start=2):
    sheet[f'A{i}'] = agent

wb.save(file_path)
print("SUCCESS")
