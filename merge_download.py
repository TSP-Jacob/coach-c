import openpyxl
import sys
from openpyxl.worksheet.table import TableColumn

file1 = r"C:\Users\jacob\Downloads\Realtor_Teams_Gatineau_Ottawa.xlsx"
file2 = r"C:\Coach-C\agents.xlsx"

try:
    with open(file2, 'a'):
        pass
except PermissionError:
    print("PERMISSION_ERROR")
    sys.exit(1)

wb1 = openpyxl.load_workbook(file1, data_only=True)
wb2 = openpyxl.load_workbook(file2)

target_sheet = wb2['Target Teams']

# Get existing headers in target sheet
headers2 = {}
for cell in target_sheet[1]:
    if cell.value:
        headers2[str(cell.value)] = cell.column

# Add missing headers from downloaded file
downloaded_headers = ['Leader / Contact', 'Brokerage', 'Notable Info', 'Contacted', 'Response', 'Notes']
max_col = target_sheet.max_column
for h in downloaded_headers:
    if h not in headers2:
        max_col += 1
        target_sheet.cell(row=1, column=max_col, value=h)
        headers2[h] = max_col

# Identify existing agents to prevent duplicates
seen_agents = set()
if 'Agents' in headers2:
    col_idx = headers2['Agents']
    for row in range(2, target_sheet.max_row + 1):
        val = target_sheet.cell(row=row, column=col_idx).value
        if val:
            seen_agents.add(str(val).strip().lower())

next_row = target_sheet.max_row + 1
added_count = 0

for sheet_name in wb1.sheetnames:
    city = "Gatineau" if "Gatineau" in sheet_name else "Ottawa"
    s1 = wb1[sheet_name]
    
    # map s1 headers
    h1 = {}
    for cell in s1[1]:
        if cell.value:
            h1[str(cell.value)] = cell.column
            
    if 'Team Name' not in h1:
        continue
        
    for row in range(2, s1.max_row + 1):
        team_name = s1.cell(row=row, column=h1['Team Name']).value
        if not team_name:
            continue
            
        clean_name = str(team_name).strip()
        lower_name = clean_name.lower()
        
        if lower_name not in seen_agents:
            seen_agents.add(lower_name)
            
            # Write to target sheet
            target_sheet.cell(row=next_row, column=headers2['Agents'], value=clean_name)
            if 'City' in headers2:
                target_sheet.cell(row=next_row, column=headers2['City'], value=city)
                
            # Map other fields
            for source_h in downloaded_headers:
                if source_h in h1:
                    val = s1.cell(row=row, column=h1[source_h]).value
                    if val is not None:
                        target_sheet.cell(row=next_row, column=headers2[source_h], value=val)
                        
            # meeting booked
            if 'Meeting Booked' in h1 and 'Meeting' in headers2:
                val = s1.cell(row=row, column=h1['Meeting Booked']).value
                if val is not None:
                    target_sheet.cell(row=next_row, column=headers2['Meeting'], value=val)
                    
            next_row += 1
            added_count += 1

# Update table
if target_sheet.tables:
    table = list(target_sheet.tables.values())[0]
    existing_cols = len(table.tableColumns)
    
    if existing_cols < max_col:
        for col_idx in range(existing_cols + 1, max_col + 1):
            header_name = str(target_sheet.cell(row=1, column=col_idx).value)
            new_col = TableColumn(id=col_idx, name=header_name)
            table.tableColumns.append(new_col)
            
    table.ref = f"A1:{openpyxl.utils.get_column_letter(max_col)}{target_sheet.max_row}"

wb2.save(file2)
print(f"SUCCESS: Added {added_count} new teams.")
