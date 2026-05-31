import openpyxl

file1 = r"C:\Users\jacob\Downloads\Realtor_Teams_Gatineau_Ottawa.xlsx"
file2 = r"C:\Coach-C\agents.xlsx"

try:
    wb1 = openpyxl.load_workbook(file1)
    print(f"Downloaded file sheets: {wb1.sheetnames}")
    s1 = wb1.active
    headers1 = [cell.value for cell in s1[1]]
    print(f"Downloaded file headers: {headers1}")
    print("Downloaded file first data row:")
    print([cell.value for cell in s1[2]])
except Exception as e:
    print(f"Error reading downloaded file: {e}")

try:
    wb2 = openpyxl.load_workbook(file2)
    s2 = wb2['Target Teams']
    headers2 = [cell.value for cell in s2[1]]
    print(f"\nagents.xlsx Target Teams headers: {headers2}")
except Exception as e:
    print(f"Error reading agents.xlsx: {e}")
