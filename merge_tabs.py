import openpyxl
import sys

file_path = 'agents.xlsx'

try:
    with open(file_path, 'a'):
        pass
except PermissionError:
    print("PERMISSION_ERROR")
    sys.exit(1)

wb = openpyxl.load_workbook(file_path)

if 'Gatineau' not in wb.sheetnames:
    print("Gatineau tab not found!")
    sys.exit(1)

gatineau = wb['Gatineau']

# 1. Rename Column1 to City in row 1
city_col_idx = 4 # Column D is index 4 (1-based), but let's find it dynamically or just use D
for cell in gatineau[1]:
    if cell.value == 'Column1':
        city_col_idx = cell.column
        cell.value = 'City'
        break
else:
    # If not found, just use column 4
    gatineau.cell(row=1, column=4, value='City')
    city_col_idx = 4

# Update table column name
if gatineau.tables:
    table = list(gatineau.tables.values())[0]
    # tableColumns is 0-indexed
    if len(table.tableColumns) >= city_col_idx:
        table.tableColumns[city_col_idx - 1].name = 'City'

# 2. Gather existing agents in Gatineau
seen_names = set()
for row in range(2, gatineau.max_row + 1):
    agent_name = gatineau.cell(row=row, column=1).value
    if agent_name:
        clean_name = str(agent_name).strip().lower()
        seen_names.add(clean_name)
        # Set city to Gatineau for existing rows if empty
        if not gatineau.cell(row=row, column=city_col_idx).value:
            gatineau.cell(row=row, column=city_col_idx, value='Gatineau')

next_row = gatineau.max_row + 1

# 3. Iterate through other sheets
sheets_to_delete = []
for sheet_name in wb.sheetnames:
    if sheet_name == 'Gatineau':
        continue
        
    sheet = wb[sheet_name]
    
    # Assume agents are in column A
    for row in range(2, sheet.max_row + 1):
        agent_name = sheet.cell(row=row, column=1).value
        if agent_name:
            clean_name = str(agent_name).strip()
            lower_name = clean_name.lower()
            if lower_name not in seen_names:
                seen_names.add(lower_name)
                # Add to Gatineau sheet
                gatineau.cell(row=next_row, column=1, value=clean_name)
                gatineau.cell(row=next_row, column=city_col_idx, value=sheet_name)
                next_row += 1
                
    sheets_to_delete.append(sheet_name)

# 4. Delete other sheets
for sheet_name in sheets_to_delete:
    del wb[sheet_name]

# 5. Rename Gatineau to Quebec
gatineau.title = 'Quebec'

# 6. Update table ref
if gatineau.tables:
    table = list(gatineau.tables.values())[0]
    table.ref = f"A1:K{next_row - 1}"

wb.save(file_path)
print("SUCCESS")
